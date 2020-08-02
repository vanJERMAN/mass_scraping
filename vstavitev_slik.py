import pandas as pd
# import xlsxwriter
import openpyxl
from openpyxl.styles import Alignment
from datetime import date
from posiljanje import posiljanje

try:
	def vstavitev_slike():
		datum = date.today()
		data = pd.read_excel(rf'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx')
		df = pd.DataFrame(data, columns=["SLIKA"])


		wb = openpyxl.load_workbook(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")
		ws = wb.worksheets[0] 
		sheet = wb.active
		ws.column_dimensions['A'].width = 15
		ws.column_dimensions['B'].width = 38
		ws.column_dimensions['C'].width = 50
		column_sizes = ("K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U")
		for s in column_sizes:
			ws.column_dimensions[s].width = 80
			ws.column_dimensions[s].height = 50
		# workbook = xlsxwriter.Workbook('/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/mass_data copy.xlsx')
		# worksheet = workbook.add_worksheet("ola")

		# worksheet.set_column('C1:C5', 70)
		# worksheet.set_default_row(45)

		# image_row = 1
		# image_col = 0 
		i = 2
		row = 2
		for g in range(1, sheet.max_column):
			sheet.cell(1, g).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		#spremeni da je na len() konec
		rows = range(1, sheet.max_row + 1)
		columns = range(2, sheet.max_column + 1)

		for d in df["SLIKA"]:

			sheet.row_dimensions[row].height = 150
			img = openpyxl.drawing.image.Image(d)
			img.anchor = f"A{i}"
			ws.add_image(img)


			i += 1
			row += 1

		for row in range(2, sheet.max_row+1):
		    for col in range(2 ,11):
		        sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		for row in range(2, sheet.max_row+1):
		    for col in range(11, sheet.max_column+1):
		        sheet.cell(row, col).alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)



		# 	image = d
		# 	worksheet.insert_image(image_row, image_col, image, {"x_scale": 0.12, "y_scale": 0.12, "x_offset": 15, "y_offset": 5, "positioning": 1})
		# 	image_row += 1

		# workbook.close()
		wb.save(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")
		# wb.close()

	if __name__ == "__main__":
		vstavitev_slike()



except:
	try:
		print("Trying again to run vstavitev_slik.py")
		vstavitev_slike()
	except Exception as error:
		posiljanje("vanjermancek@gmail.com", "There was an error while running vstavitev_slik.py", f"{error}")
