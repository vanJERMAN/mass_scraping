import matplotlib.pyplot as plt
import xlrd
import os
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Alignment
import pandas as pd
from collections import defaultdict
from time import sleep
from posiljanje import posiljanje

try:
	def izdelava_grafov():
		vsi_artikli_na_zalogi = 0
		vsi_artikli_ni_na_zalogi = 0

		plt.style.use('ggplot')


		plt.rc('axes', titlesize=8)
		plt.rc('xtick', labelsize=7)
		fig2 = plt.figure(figsize=(3.0, 2.0))
		datum = date.today()
		wb = openpyxl.load_workbook(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")
		ws = wb.worksheets[0] 
		sheet = wb.active
		path = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx"

		for row in range(1, sheet.max_row):
			celoten_artikel_na_zalogi = 0
			celoten_artikel_ni_na_zalogi = 0


			for st in range(10, sheet.max_column):
				na_zalogi=0
				ni_na_zalogi=0
				inputWorkbook = xlrd.open_workbook(path)
				inputWorksheet = inputWorkbook.sheet_by_index(0)
				splitano = inputWorksheet.cell_value(row, st).replace(": ", "\n").split("\n")
				brand = inputWorksheet.cell_value(row, 3)
				ime = inputWorksheet.cell_value(row, 4)
				stevilka = inputWorksheet.cell_value(0, st)[-5:]
				url_index = inputWorksheet.cell_value(row, 9)[-22:-4]
				# print(splitano)
				# print(len(splitano))
				index = [x for x in range(1, len(splitano)) if x % 2 == 1]

				for i in index:
					if splitano[i][0:9] == "NA ZALOGI":
						na_zalogi += 1
						celoten_artikel_na_zalogi += 1
						vsi_artikli_na_zalogi += 1
					else:
						ni_na_zalogi += 1
						celoten_artikel_ni_na_zalogi += 1
						vsi_artikli_ni_na_zalogi += 1
					i += 1

					# print(sheet.max_column-1)
					# print(ni_na_zalogi)

				if na_zalogi > 0 and ni_na_zalogi > 0:	
					

					slices = [na_zalogi, ni_na_zalogi]
					labels = [f"Na zalogi: {na_zalogi} trgovin", f"Ni na zalogi: {ni_na_zalogi} trgovin"]
					colors = ["#e25d5d", "#a10000"]
					explode = [0.1, 0]
					# plt.style.use("grayscale")
					plt.pie(slices, explode, colors=colors, wedgeprops={"edgecolor": "black"})
					plt.title(f"Zaloga artikla {brand} {ime}, {stevilka}")
					plt.legend(labels)
					plt.tight_layout()
					# plt.draw() 
					# plt.pause(0.02)
					# plt.show()
					datum = date.today()
					# print(datum)
					plt.savefig(f'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/{url_index}.{stevilka}.{datum}.png')
					print(f'{url_index}.{stevilka}.{datum}.png')
					plt.clf()
					plt.cla()
				else:
					print("Artikel ne obstaja")

				if st == sheet.max_column-1:

					if celoten_artikel_na_zalogi == 0:
						celoten_artikel_ni_na_zalogi += 1
						slices = [celoten_artikel_na_zalogi, celoten_artikel_ni_na_zalogi]
						labels = ["Na zalogi", "Ni več na zalogi!"]
						colors = ["#4567c7", "#012380"]
						explode = [0.1, 0]
						# plt.style.use("grayscale")
						plt.pie(slices, explode, colors=colors, wedgeprops={"edgecolor": "black"})
						plt.title(f"Zaloga artikla {brand} {ime}, vse številke")
						plt.legend(labels)
						plt.tight_layout()
						datum = date.today()
						plt.savefig(f'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/{url_index}.vse_stevilke.{datum}.png')
						print(f'{url_index}.vse_stevilke.{datum}.png')
						plt.clf()
						plt.cla()
						celoten_artikel_ni_na_zalogi -= 1
						ws[f'B{row+1}'] = f"{celoten_artikel_na_zalogi}:{celoten_artikel_ni_na_zalogi}"

					elif celoten_artikel_na_zalogi > 0:
						slices = [celoten_artikel_na_zalogi, celoten_artikel_ni_na_zalogi]
						ws[f'B{row+1}'] = f"{celoten_artikel_na_zalogi}:{celoten_artikel_ni_na_zalogi}"
						print(row+1)
						labels = [f"Na zalogi: {celoten_artikel_na_zalogi} trgovin", f"Ni na zalogi: {celoten_artikel_ni_na_zalogi} trgovin"]
						colors = ["#4567c7", "#012380"]
						explode = [0.1, 0]
						# plt.style.use("grayscale")
						plt.pie(slices, explode, colors=colors, wedgeprops={"edgecolor": "black"})
						plt.title(f"Zaloga artikla {brand} {ime}, vse številke")
						plt.legend(labels)
						plt.tight_layout()
						plt.savefig(f'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/{url_index}.vse_stevilke.{datum}.png')
						print(f'{url_index}.vse_stevilke.{datum}.png')
						plt.clf()
						plt.cla()
						# print(celoten_artikel_na_zalogi)
						# print(celoten_artikel_ni_na_zalogi)
					else:
						print("Nekaj je narobe z: celoten_artikel_na_zalogi")

				if row == sheet.max_row-1:
					slices = [vsi_artikli_na_zalogi, vsi_artikli_ni_na_zalogi]
					ws[f'B{row+2}'] = f"{vsi_artikli_na_zalogi}:{vsi_artikli_ni_na_zalogi}"
					labels = [f"Na zalogi: {vsi_artikli_na_zalogi} trgovin", f"Ni na zalogi: {vsi_artikli_ni_na_zalogi} trgovin"]
					colors = ["#626262", "#000000"]
					explode = [0.1, 0]
					# plt.style.use("grayscale")
					plt.pie(slices, explode, colors=colors, wedgeprops={"edgecolor": "black"})
					plt.title(f"Vsi artikli")
					plt.tight_layout()
					plt.legend(labels)
					plt.savefig(f'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/vsi_artikli.{datum}.png')
					print(f'vsi_artikli.{datum}.png')
					plt.clf()
					plt.cla()

		wb.save(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")





	def izdelava_stolpcev():
		zaloga_izdelkov = defaultdict(list)
		zaloga_vseh_izdelkov = defaultdict(list)
		erik = pd.date_range(end = datetime.today(), periods = 36).to_pydatetime().tolist()
		datumi=[]
		plt.style.use('ggplot')
		plt.rc('axes', titlesize=3)
		plt.rc('xtick', labelsize=6)
		plt.rc('ytick', labelsize=6)
		fig2 = plt.figure(figsize=(4.0, 2.0))

		for x in range (0, 36):
			datumi.append(str(erik[x])[:10])
		# print(datumi)
		for datum in datumi:
			print(datum)
			try:
				wb = openpyxl.load_workbook(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")
				ws = wb.worksheets[0]
				sheet = wb.active
				zaloga_izdelkov["datum"].append(datum)
				zaloga_vseh_izdelkov["datum"].append(datum)
				zgodovina_vseh_izdelkov = ws[f"B{sheet.max_row}"].value
				zgodovina_vseh_izdelkov = zgodovina_vseh_izdelkov.split(":")
				zgodovina_vseh_izdelkov = zgodovina_vseh_izdelkov[0]
				zaloga_vseh_izdelkov["vsi"].append(zgodovina_vseh_izdelkov)
				for row in range(2, sheet.max_row+1):
					path = ws[f"A{row}"].value
					values = ws[f"B{row}"].value
					value = values.split(":")
					na_zalogi_zgodovina = value[0]
					zaloga_izdelkov[path].append(na_zalogi_zgodovina)
			except (FileNotFoundError, AttributeError):
				None


		for row in range(2, sheet.max_row+1):
			print(row)
			print(sheet.max_row+1)
			path_xlsx = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx"
			inputWorkbook = xlrd.open_workbook(path_xlsx)
			inputWorksheet = inputWorkbook.sheet_by_index(0)
			url_index = inputWorksheet.cell_value(row-1, 9)[-22:-4]
			path = ws[f"A{row}"].value
			brand = ws[f"D{row}"].value
			artikel = ws[f"E{row}"].value
			values = ws[f"B{row}"].value
			plt.style.use("fivethirtyeight")
			ages_x = []
			dev_y = []
			for x in range(1, 31):	
				try:
					dev_y.append(int(zaloga_izdelkov[path][-x]))
				except (IndexError, ValueError):
					None
			for y in range(1, len(dev_y)+1):
				ages_x.append(zaloga_izdelkov["datum"][-y][5:10])
				print(zaloga_izdelkov["datum"][-y][5:10])


			plt.bar(ages_x[::-1], dev_y[::-1], color="#4567c7", label="Število artiklov")
			# print(ages_x)
			# print(dev_y)
			# plt.legend()
			plt.title(f"Zgodovina prodaje {brand} {artikel}", fontdict = {'fontsize' : 12})
			# plt.xlabel("Datum")
			# plt.ylabel("Artiklov na zalogi")
			# print(ages_x)
			# print(dev_y)
			plt.xticks(rotation=90)
			plt.tight_layout()
			plt.savefig(f'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/zgodovina.{url_index}.{datum}.png')
			print(f'zgodovina.{url_index}.{datum}.png graf je shranjen!')
			plt.clf()
			plt.cla()
		plt.style.use("fivethirtyeight")
		ages_x = []
		dev_y = []
		for x in range(1, 31):	
			try:
				dev_y.append(int(zaloga_vseh_izdelkov["vsi"][-x]))
			except (IndexError, ValueError):
				None
		for y in range(1, len(dev_y)+1):
			ages_x.append(zaloga_izdelkov["datum"][-y][5:10])
			print(zaloga_izdelkov["datum"][-y][5:10])


		plt.bar(ages_x[::-1], dev_y[::-1], color="#000000", label="Število vseh artiklov")
			# print(ages_x)
			# print(dev_y)
			# plt.legend()
		plt.title(f"Zgodovina prodaje vseh artiklov", fontdict = {'fontsize' : 12})
			# plt.xlabel("Datum")
			# plt.ylabel("Artiklov na zalogi")
			# print(ages_x)
			# print(dev_y)
		plt.xticks(rotation=90)
		plt.tight_layout()
		plt.savefig(f'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/zgodovina_prodaje_vseh_artiklov.{datum}.png')
		print(f'zgodovina_prodaje_vseh_artiklov.{datum}.png graf je shranjen!')
		plt.clf()
		plt.cla()
		# for row in range(2, sheet.max_row+1):
		# 			path = ws[f"A{row}"].value
		# 			brand = ws[f"D{row}"].value
		# 			artikel = ws[f"E{row}"].value
		# 			plt.style.use("fivethirtyeight")

		# 			ages_x = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100]
		# 			dev_y = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100]

		# 			plt.bar(ages_x, dev_y, color="#4567c7", label="All Devs")

		# 			plt.legend()
		# 			plt.title(f"Zgodovina prodaje {brand} {artikel}")
		# 			plt.xlabel("Datum")
		# 			plt.ylabel("Artiklov na zalogi")
		# 			plt.tight_layout()
		# 			plt.show()
		# 			sleep(5)







	def vstavitev_grafov():
		datum = date.today()
		wb = openpyxl.load_workbook(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")
		ws = wb.worksheets[0] 
		sheet = wb.active
		path = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx"

		inputWorkbook = xlrd.open_workbook(path)
		inputWorksheet = inputWorkbook.sheet_by_index(0)
		datum = date.today()
		data = pd.read_excel(rf'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx')
		img = openpyxl.drawing.image.Image(f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/vsi_artikli.{datum}.png")	
		ws.add_image(img, f"B{sheet.max_row}")
		print(f"Celotna kolekcija graf je dodan")
		img = openpyxl.drawing.image.Image(f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/zgodovina_prodaje_vseh_artiklov.{datum}.png")	
		ws.add_image(img, f"C{sheet.max_row}")
		print(f"Celotna zgodovina graf je dodan")

		for row in range(1, sheet.max_row-1):
			inputWorkbook = xlrd.open_workbook(path)
			inputWorksheet = inputWorkbook.sheet_by_index(0)
			url_index = inputWorksheet.cell_value(row, 9)[-22:-4]
			datum = date.today()
			data = pd.read_excel(rf'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx')
			img = openpyxl.drawing.image.Image(f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/{url_index}.vse_stevilke.{datum}.png")	
			ws.add_image(img, f"B{row+1}")
			print(f"row {row} je dodana slika")
			img = openpyxl.drawing.image.Image(f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/zgodovina.{url_index}.{datum}.png")	
			ws.add_image(img, f"C{row+1}")
			print(f"Zgodovina {row} je dodana slika")


			for st in range(10, sheet.max_column):
				inputWorkbook = xlrd.open_workbook(path)
				inputWorksheet = inputWorkbook.sheet_by_index(0)
				stevilka = inputWorksheet.cell_value(0, st)[-5:]
				url_index = inputWorksheet.cell_value(row, 9)[-22:-4]
				datum = date.today()
				data = pd.read_excel(rf'/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx')

				try:
					img = openpyxl.drawing.image.Image(f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/slike/grafi/{url_index}.{stevilka}.{datum}.png")
					if st == 10:
						ws.add_image(img, f"K{row+1}")
						print("36")
					elif st == 11:
						ws.add_image(img, f"L{row+1}")
						print("37")
					elif st == 12:
						ws.add_image(img, f"M{row+1}")
						print("38")
					elif st == 13:
						ws.add_image(img, f"N{row+1}")
						print("39")
					elif st == 14:
						ws.add_image(img, f"O{row+1}")
						print("40")
					elif st == 15:
						ws.add_image(img, f"P{row+1}")
						print("41")
					elif st == 16:
						ws.add_image(img, f"Q{row+1}")
						print("42")
					elif st == 17:
						ws.add_image(img, f"R{row+1}")
						print("43")
					elif st == 18:
						ws.add_image(img, f"S{row+1}")
						print("44")
					elif st == 19:
						ws.add_image(img, f"T{row+1}")
						print("45")
					elif st == 20:
						ws.add_image(img, f"U{row+1}")
						print("46")
					else:
						print("Nekai je narobe pri st")
					# else:
					# 	print("Ni številke")
				except FileNotFoundError:
					print("Ni te št")



		wb.save(filename = f"/Users/erik/Desktop/PYTHON/CSV/beautiful_soup/csv_in_xlsx_datoteke/{datum}.xlsx")



	if __name__ == "__main__":
		izdelava_grafov()
		izdelava_stolpcev()
		vstavitev_grafov()
except:
	try:
		print("Trying again to run grafi.py")
		izdelava_grafov()
		izdelava_stolpcev()
		vstavitev_grafov()
	except Exception as error:
		posiljanje("vanjermancek@gmail.com", "There was an error while running grafi.py", f"{error}")

